"""
peer.py — Peer Comparison Engine
Sprint 3: Intra-group percentile ranking, radar chart data,
side-by-side comparison tables, best-in-class and watch-list detection.

Usage:
    python src/analytics/peer.py
    OR imported by dashboard and API.
"""

import sqlite3
import logging
from pathlib import Path

import pandas as pd
import numpy as np

BASE_DIR   = Path(__file__).resolve().parents[2]
DB_PATH    = BASE_DIR / "data" / "nifty100.db"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger(__name__)

# 8 metrics used for radar charts (must exist in computed_ratios)
RADAR_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
]

# 10 core metrics for percentile ranking
RANK_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "health_score",
]

# Metrics where LOWER is better (inverted for percentile rank)
LOWER_IS_BETTER = {"debt_to_equity"}


# ─────────────────────────────────────────────────────────────────────────────
# Data loader
# ─────────────────────────────────────────────────────────────────────────────

def load_peer_universe(conn: sqlite3.Connection) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load latest-year computed_ratios and peer_groups tables.
    Returns (ratios_df, peer_groups_df).
    """
    ratios = pd.read_sql("""
        SELECT cr.company_id, cr.year, c.company_name,
               cr.return_on_equity_pct, cr.return_on_capital_pct,
               cr.net_profit_margin_pct, cr.operating_profit_margin_pct,
               cr.ebit_margin_pct, cr.debt_to_equity, cr.interest_coverage,
               cr.is_debt_free, cr.free_cash_flow_cr, cr.cfo_to_pat_ratio,
               cr.revenue_cagr_3yr, cr.revenue_cagr_5yr, cr.revenue_cagr_10yr,
               cr.pat_cagr_3yr, cr.pat_cagr_5yr, cr.eps_cagr_5yr,
               cr.asset_turnover, cr.book_value_per_share,
               cr.earnings_per_share, cr.health_score, cr.health_band,
               cr.capital_alloc_pattern
        FROM computed_ratios cr
        JOIN companies c ON cr.company_id = c.id
        WHERE cr.year = (
            SELECT MAX(year) FROM computed_ratios cr2
            WHERE cr2.company_id = cr.company_id
        )
    """, conn)

    peer_groups = pd.read_sql(
        "SELECT peer_group_name, company_id, is_benchmark FROM peer_groups",
        conn
    )

    # Winsorise ROE cap to 200%
    ratios["return_on_equity_pct"] = ratios["return_on_equity_pct"].clip(upper=200.0)

    logger.info("Peer universe: %d companies, %d peer groups",
                len(ratios), peer_groups["peer_group_name"].nunique())
    return ratios, peer_groups


# ─────────────────────────────────────────────────────────────────────────────
# Percentile ranking per group
# ─────────────────────────────────────────────────────────────────────────────

def compute_peer_percentiles(
    ratios: pd.DataFrame,
    peer_groups: pd.DataFrame
) -> pd.DataFrame:
    """
    For each peer group × metric, compute PERCENT_RANK (0–1).
    For lower-is-better metrics (D/E), rank is inverted so
    lower value = higher percentile.
    Returns long-format DataFrame with columns:
        peer_group_name, company_id, metric, value, percentile_rank, year
    """
    # Join ratios with peer group membership
    joined = peer_groups.merge(ratios, on="company_id", how="left")

    rows = []
    for group_name, grp_df in joined.groupby("peer_group_name"):
        grp_df = grp_df.dropna(subset=["company_id"])
        n = len(grp_df)
        if n < 2:
            continue

        for metric in RANK_METRICS:
            if metric not in grp_df.columns:
                continue
            vals = pd.to_numeric(grp_df[metric], errors="coerce")
            ascending = metric not in LOWER_IS_BETTER

            # PERCENT_RANK: 0 = lowest, 1 = highest
            ranked = vals.rank(method="average", ascending=ascending, na_option="bottom")
            pct_rank = ((ranked - 1) / max(n - 1, 1)).round(4)

            for _, row in grp_df.iterrows():
                val = vals.get(row.name)
                pct = pct_rank.get(row.name)
                rows.append({
                    "peer_group_name": group_name,
                    "company_id":      row["company_id"],
                    "is_benchmark":    row["is_benchmark"],
                    "metric":          metric,
                    "value":           val,
                    "percentile_rank": pct,
                    "year":            row.get("year"),
                })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Best-in-class and watch-list detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_best_and_weak(percentiles_df: pd.DataFrame) -> pd.DataFrame:
    """
    Best-in-class: top quartile (≥75th pctile) for ≥6 of 10 metrics.
    Watch-list: bottom quartile (≤25th pctile) for ≥4 of 10 metrics.
    Returns DataFrame with company_id, peer_group_name, badge.
    """
    rows = []
    for (group, company), grp in percentiles_df.groupby(["peer_group_name", "company_id"]):
        top_q    = (grp["percentile_rank"] >= 0.75).sum()
        bottom_q = (grp["percentile_rank"] <= 0.25).sum()

        badge = "Standard"
        if top_q >= 6:
            badge = "Best in Class"
        elif bottom_q >= 4:
            badge = "Watch List"

        rows.append({
            "peer_group_name": group,
            "company_id":      company,
            "top_quartile_count":    int(top_q),
            "bottom_quartile_count": int(bottom_q),
            "badge":           badge,
        })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Radar chart data
# ─────────────────────────────────────────────────────────────────────────────

def build_radar_data(
    ratios: pd.DataFrame,
    peer_groups: pd.DataFrame
) -> dict[str, dict]:
    """
    For each (company, peer_group) pair, build radar chart data:
    8 normalised axes (0–100) for the company and the group average.
    Returns dict: company_id -> {group_name, axes: [{metric, company_val,
                                 group_avg, normalised}]}
    """
    joined = peer_groups.merge(ratios, on="company_id", how="left")
    radar_data = {}

    for group_name, grp_df in joined.groupby("peer_group_name"):
        # Compute group average for each radar metric
        group_avgs = {}
        for m in RADAR_METRICS:
            if m in grp_df.columns:
                vals = pd.to_numeric(grp_df[m], errors="coerce")
                group_avgs[m] = vals.mean()

        # Min-max normalise across group for radar
        group_ranges = {}
        for m in RADAR_METRICS:
            if m in grp_df.columns:
                vals = pd.to_numeric(grp_df[m], errors="coerce")
                mn, mx = vals.min(), vals.max()
                group_ranges[m] = (mn, mx)

        for _, row in grp_df.iterrows():
            company_id = row["company_id"]
            axes = []
            for m in RADAR_METRICS:
                if m not in grp_df.columns:
                    continue
                raw_val = pd.to_numeric(row.get(m), errors="coerce")
                avg_val = group_avgs.get(m)
                mn, mx  = group_ranges.get(m, (0, 1))

                # Normalise 0–100
                if mx != mn and not pd.isna(raw_val):
                    if m in LOWER_IS_BETTER:
                        norm = (1 - (raw_val - mn) / (mx - mn)) * 100
                    else:
                        norm = (raw_val - mn) / (mx - mn) * 100
                else:
                    norm = 50.0

                axes.append({
                    "metric":       m,
                    "company_val":  round(float(raw_val), 2) if not pd.isna(raw_val) else None,
                    "group_avg":    round(float(avg_val), 2) if avg_val is not None else None,
                    "normalised":   round(float(np.clip(norm, 0, 100)), 1),
                })

            key = f"{company_id}::{group_name}"
            radar_data[key] = {
                "company_id":      company_id,
                "peer_group_name": group_name,
                "axes":            axes,
            }

    return radar_data


# ─────────────────────────────────────────────────────────────────────────────
# Side-by-side comparison table
# ─────────────────────────────────────────────────────────────────────────────

COMPARISON_METRICS = [
    "return_on_equity_pct", "return_on_capital_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage",
    "free_cash_flow_cr", "cfo_to_pat_ratio", "revenue_cagr_3yr",
    "revenue_cagr_5yr", "pat_cagr_3yr", "pat_cagr_5yr",
    "earnings_per_share", "book_value_per_share", "health_score",
]


def build_comparison_table(
    company_ids: list[str],
    ratios: pd.DataFrame
) -> pd.DataFrame:
    """
    Side-by-side metric table for up to 5 companies.
    Returns DataFrame: metric × company_id.
    """
    subset = ratios[ratios["company_id"].isin(company_ids)].set_index("company_id")
    available = [m for m in COMPARISON_METRICS if m in subset.columns]
    table = subset[available].T
    return table.round(2)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

def export_peer_comparison(
    ratios: pd.DataFrame,
    peer_groups: pd.DataFrame,
    percentiles: pd.DataFrame,
    badges: pd.DataFrame,
    output_path: Path,
) -> None:
    """Export peer_comparison.xlsx with one sheet per peer group."""
    try:
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        HAS_STYLE = True
    except ImportError:
        HAS_STYLE = False

    joined = peer_groups.merge(ratios, on="company_id", how="left")
    joined = joined.merge(
        badges[["peer_group_name", "company_id", "badge"]],
        on=["peer_group_name", "company_id"], how="left"
    )

    # Merge percentile pivot (wide format per group)
    pct_pivot = (
        percentiles
        .pivot_table(index=["peer_group_name", "company_id"],
                     columns="metric", values="percentile_rank")
        .reset_index()
    )
    pct_pivot.columns = [
        f"pctile_{c}" if c not in ("peer_group_name", "company_id") else c
        for c in pct_pivot.columns
    ]

    joined = joined.merge(pct_pivot, on=["peer_group_name", "company_id"], how="left")

    # Drop any duplicate columns introduced by merges
    joined = joined.loc[:, ~joined.columns.duplicated()]

    display_cols = (
        ["company_id", "company_name", "badge", "is_benchmark", "health_score", "health_band"]
        + COMPARISON_METRICS
        + [c for c in joined.columns if c.startswith("pctile_")]
    )
    display_cols = [c for c in display_cols if c in joined.columns]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        written = 0
        for group_name, grp in joined.groupby("peer_group_name"):
            sheet_df = grp[display_cols].copy()
            # Remove any duplicate columns from merge artifacts
            sheet_df = sheet_df.loc[:, ~sheet_df.columns.duplicated()]

            # Sort: benchmark first, then by health_score desc
            sort_cols = [c for c in ["is_benchmark", "health_score"]
                         if c in sheet_df.columns]
            if sort_cols:
                sheet_df = sheet_df.sort_values(
                    sort_cols, ascending=[False] * len(sort_cols)
                )

            for col in sheet_df.select_dtypes(include="number").columns:
                sheet_df[col] = sheet_df[col].round(2)

            sheet_name = group_name[:31]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            written += 1

            if HAS_STYLE:
                ws = writer.sheets[sheet_name]
                # Colour-code percentile columns: green ≥0.75, red ≤0.25
                pctile_cols = [
                    c for c in display_cols if c.startswith("pctile_")
                ]
                for col_idx, col_name in enumerate(display_cols, 1):
                    if col_name in pctile_cols:
                        for row_idx in range(2, len(sheet_df) + 2):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            val  = cell.value
                            if val is not None:
                                try:
                                    v = float(val)
                                    if v >= 0.75:
                                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                                    elif v <= 0.25:
                                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                                except (ValueError, TypeError):
                                    pass

                # Auto column widths
                for col_cells in ws.columns:
                    max_len = max(
                        (len(str(cell.value or "")) for cell in col_cells),
                        default=10
                    )
                    ws.column_dimensions[
                        get_column_letter(col_cells[0].column)
                    ].width = min(max_len + 2, 28)

    logger.info("Peer comparison written: %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main runner
# ─────────────────────────────────────────────────────────────────────────────

def run_peer_engine() -> dict:
    """Full peer comparison pipeline."""
    import time
    t0 = time.time()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    logger.info("=" * 55)
    logger.info("PEER ENGINE — Starting")
    logger.info("=" * 55)

    conn = sqlite3.connect(str(DB_PATH))

    ratios, peer_groups = load_peer_universe(conn)

    # Percentile ranks
    percentiles = compute_peer_percentiles(ratios, peer_groups)
    logger.info("Percentiles computed: %d rows", len(percentiles))

    # Write to SQLite
    percentiles.to_sql("peer_percentiles", conn, if_exists="replace", index=False)

    # Badges
    badges = detect_best_and_weak(percentiles)
    badges.to_csv(OUTPUT_DIR / "peer_badges.csv", index=False)
    logger.info("Badges: %s", badges["badge"].value_counts().to_dict())

    # Radar data (JSON-serialisable)
    radar_data = build_radar_data(ratios, peer_groups)
    logger.info("Radar data built: %d company-group pairs", len(radar_data))

    # Excel export
    peer_xlsx = OUTPUT_DIR / "peer_comparison.xlsx"
    export_peer_comparison(ratios, peer_groups, percentiles, badges, peer_xlsx)

    conn.close()
    elapsed = round(time.time() - t0, 2)

    # Summary
    best  = badges[badges["badge"] == "Best in Class"]
    watch = badges[badges["badge"] == "Watch List"]

    print(f"\n{'='*55}")
    print("PEER ENGINE RESULTS")
    print(f"{'='*55}")
    print(f"  Peer groups processed : {peer_groups['peer_group_name'].nunique()}")
    print(f"  Companies ranked      : {peer_groups['company_id'].nunique()}")
    print(f"  Percentile rows       : {len(percentiles)}")
    print(f"  Best-in-Class badges  : {len(best)}")
    print(f"  Watch-List flags      : {len(watch)}")
    print(f"\n  Best in Class:")
    for _, r in best.iterrows():
        print(f"    {r['company_id']:<15} ({r['peer_group_name']})")
    print(f"\n  Watch List:")
    for _, r in watch.iterrows():
        print(f"    {r['company_id']:<15} ({r['peer_group_name']})")
    print(f"\n  peer_comparison.xlsx  : {peer_xlsx}")
    print(f"  Runtime               : {elapsed}s")

    return {
        "percentiles": percentiles,
        "badges":      badges,
        "radar_data":  radar_data,
    }


if __name__ == "__main__":
    run_peer_engine()
